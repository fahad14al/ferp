"""
Utility functions for report generation and export
"""
from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO
import csv
from datetime import datetime


def export_to_csv(data, filename, headers):
    """
    Export data to CSV format
    
    Args:
        data: List of dictionaries or list of lists
        filename: Name of the file to download
        headers: List of column headers
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for row in data:
        if isinstance(row, dict):
            writer.writerow([row.get(header, '') for header in headers])
        else:
            writer.writerow(row)
    
    return response


def export_to_excel(data, filename, headers, sheet_name='Report'):
    """
    Export data to Excel format using openpyxl
    
    Args:
        data: List of dictionaries or list of lists
        filename: Name of the file to download
        headers: List of column headers
        sheet_name: Name of the Excel sheet
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, row_data in enumerate(data, 2):
            if isinstance(row_data, dict):
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
            else:
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    except ImportError:
        # Fallback to CSV if openpyxl is not installed
        return export_to_csv(data, filename.replace('.xlsx', '.csv'), headers)


def export_to_pdf(html_content, filename):
    """
    Export HTML content to PDF format using WeasyPrint
    
    Args:
        html_content: HTML string to convert to PDF
        filename: Name of the file to download
    """
    try:
        from weasyprint import HTML
        
        pdf_file = BytesIO()
        HTML(string=html_content).write_pdf(pdf_file)
        pdf_file.seek(0)
        
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    except ImportError:
        # Fallback message if WeasyPrint is not installed
        response = HttpResponse(
            "PDF export requires WeasyPrint library. Please install it with: pip install weasyprint",
            content_type='text/plain'
        )
        return response


def format_currency(value):
    """Format value as currency"""
    try:
        return f"${float(value):,.2f}"
    except (ValueError, TypeError):
        return "$0.00"


def format_percentage(value):
    """Format value as percentage"""
    try:
        return f"{float(value):.2f}%"
    except (ValueError, TypeError):
        return "0.00%"


def calculate_trend(current_value, previous_value):
    """
    Calculate trend percentage between two values
    
    Returns:
        tuple: (percentage_change, is_positive)
    """
    try:
        if previous_value == 0:
            return (100.0 if current_value > 0 else 0.0, current_value > 0)
        
        change = ((current_value - previous_value) / previous_value) * 100
        return (abs(change), change >= 0)
    except (ValueError, TypeError, ZeroDivisionError):
        return (0.0, True)


def get_date_range_label(start_date, end_date):
    """Generate a human-readable date range label"""
    if start_date and end_date:
        return f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
    elif start_date:
        return f"From {start_date.strftime('%b %d, %Y')}"
    elif end_date:
        return f"Until {end_date.strftime('%b %d, %Y')}"
    return "All Time"


def prepare_chart_data(labels, datasets):
    """
    Prepare data for Chart.js
    
    Args:
        labels: List of labels for x-axis
        datasets: List of dictionaries with 'label', 'data', and optional 'backgroundColor'
    
    Returns:
        dict: Chart.js compatible data structure
    """
    return {
        'labels': labels,
        'datasets': datasets
    }


def generate_color_palette(count):
    """Generate a color palette for charts"""
    colors = [
        '#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b',
        '#858796', '#5a5c69', '#2e59d9', '#17a673', '#2c9faf'
    ]
    return [colors[i % len(colors)] for i in range(count)]
